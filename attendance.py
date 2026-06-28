# ============================================================
#  Webcam Face Attendance System
#  Uses ONLY opencv — no dlib, no deepface, no tensorflow!
#  Install: pip install opencv-python pandas openpyxl
#  Run:     py attendance.py
# ============================================================

import cv2
import numpy as np
import pandas as pd
import os
import glob
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── CONFIG ────────────────────────────────────────────────────
KNOWN_FACES_DIR = "known_faces"
EXCEL_FILE      = "attendance.xlsx"
CONFIRM_SECONDS = 2
FONT            = cv2.FONT_HERSHEY_SIMPLEX
# ─────────────────────────────────────────────────────────────


# ── STEP 1: Load Known Faces ──────────────────────────────────
def load_known_faces():
    print("\n" + "=" * 50)
    print("  Loading student photos...")
    print("=" * 50)

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    )

    image_files = glob.glob(os.path.join(KNOWN_FACES_DIR, "*"))
    image_files = [f for f in image_files
                   if f.lower().endswith((".jpg", ".jpeg", ".png"))]

    if not image_files:
        print(f"\n  No photos found in '{KNOWN_FACES_DIR}' folder!")
        print(f"  Add photos like: Ravi.jpg  Priya.jpg")
        input("  Press Enter to exit...")
        exit()

    faces_data  = []
    labels      = []
    label_names = {}   # id -> name
    label_id    = 0

    # Group by name (supports Ravi_1.jpg, Ravi_2.jpg)
    name_files = {}
    for filepath in image_files:
        base = os.path.splitext(os.path.basename(filepath))[0]
        name = base.split("_")[0].strip()
        name_files.setdefault(name, []).append(filepath)

    for name, files in name_files.items():
        label_names[label_id] = name
        for filepath in files:
            img  = cv2.imread(filepath)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            detected = face_cascade.detectMultiScale(
                gray, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50)
            )
            if len(detected) > 0:
                x, y, w, h = detected[0]
                face_roi = cv2.resize(gray[y:y+h, x:x+w], (200, 200))
                faces_data.append(face_roi)
                labels.append(label_id)
                print(f"  ✅  {name}")
            else:
                # Use full image if no face detected in photo
                gray_resized = cv2.resize(gray, (200, 200))
                faces_data.append(gray_resized)
                labels.append(label_id)
                print(f"  ⚠️  {name} (no face detected in photo, using full image)")
        label_id += 1

    recognizer.train(faces_data, np.array(labels))
    print(f"\n  {len(label_names)} student(s) loaded. Ready!\n")
    return recognizer, face_cascade, label_names


# ── STEP 2: Excel Helpers ─────────────────────────────────────
def get_today():
    return datetime.now().strftime("%d-%b-%Y")


def load_or_create_excel(all_names):
    today = get_today()
    if os.path.exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE, index_col=0)
        for name in all_names:
            if name not in df.index:
                df.loc[name] = ""
        if today not in df.columns:
            df[today] = "Absent"
    else:
        df = pd.DataFrame(index=sorted(all_names))
        df.index.name = "Name"
        df[today] = "Absent"
    return df, today


def save_excel(df):
    df.to_excel(EXCEL_FILE)
    apply_style()
    print(f"  Saved -> {EXCEL_FILE}")


def apply_style():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    header = PatternFill("solid", fgColor="4472C4")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.fill = header
                cell.font = Font(bold=True, color="FFFFFF")
            elif cell.value == "Present":
                cell.fill = green
                cell.font = Font(color="276221")
            elif cell.value == "Absent":
                cell.fill = red
                cell.font = Font(color="9C0006")
    ws.column_dimensions["A"].width = 20
    for col in ws.columns:
        if col[0].column_letter != "A":
            ws.column_dimensions[col[0].column_letter].width = 14
    wb.save(EXCEL_FILE)


# ── STEP 3: Draw UI ───────────────────────────────────────────
def draw_box(frame, x1, y1, x2, y2, name, confidence, status, progress=0):
    if name == "Unknown":
        color = (0, 0, 220)
    elif status == "marked":
        color = (0, 180, 0)
    else:
        color = (0, 200, 255)

    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

    label = f"{name} {confidence}%" if name != "Unknown" else "Unknown"
    (w, h), _ = cv2.getTextSize(label, FONT, 0.55, 1)
    cv2.rectangle(frame, (x1, y1 - h - 8), (x1 + w + 8, y1), color, -1)
    cv2.putText(frame, label, (x1 + 4, y1 - 4),
                FONT, 0.55, (255, 255, 255), 1, cv2.LINE_AA)

    if status == "confirming" and progress > 0:
        bar_w  = x2 - x1
        filled = int(bar_w * progress)
        cv2.rectangle(frame, (x1, y2 + 4), (x2, y2 + 10), (60, 60, 60), -1)
        if filled > 0:
            cv2.rectangle(frame, (x1, y2 + 4),
                          (x1 + filled, y2 + 10), color, -1)

    if status == "marked":
        cv2.putText(frame, "Marked Present",
                    (x1, y2 + 20), FONT, 0.45,
                    (0, 180, 0), 1, cv2.LINE_AA)


def draw_hud(frame, today, present, total):
    absent = total - present
    pct    = round(present / total * 100 if total else 0)
    cv2.rectangle(frame, (0, 0), (370, 85), (20, 20, 20), -1)
    cv2.putText(frame, f"Date    : {today}",
                (10, 22), FONT, 0.52, (200, 200, 200), 1)
    cv2.putText(frame,
                f"Total: {total}  Present: {present}  Absent: {absent}",
                (10, 46), FONT, 0.52, (200, 200, 200), 1)
    cv2.putText(frame, f"Attendance : {pct}%",
                (10, 70), FONT, 0.52, (100, 255, 100), 1)
    cv2.putText(frame, "Press Q to quit & save",
                (frame.shape[1] - 215, 24),
                FONT, 0.48, (150, 150, 150), 1)


# ── STEP 4: Main Loop ─────────────────────────────────────────
def run():
    recognizer, face_cascade, label_names = load_known_faces()
    all_names = list(label_names.values())
    df, today = load_or_create_excel(all_names)

    marked_today  = set()
    confirm_start = {}

    for name in all_names:
        if today in df.columns and df.at[name, today] == "Present":
            marked_today.add(name)

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("\n  Cannot open webcam!")
        input("  Press Enter to exit...")
        return

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

    print("=" * 50)
    print("  FACE ATTENDANCE SYSTEM")
    print(f"  Date     : {today}")
    print(f"  Students : {len(all_names)}")
    print("  Press Q to quit and save")
    print("=" * 50 + "\n")

    CONFIDENCE_THRESHOLD = 80   # lower = stricter (0-100)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(
            gray, scaleFactor=1.1, minNeighbors=5, minSize=(80, 80)
        )

        now = time.time()

        for (x, y, w, h) in faces:
            x1, y1, x2, y2 = x, y, x + w, y + h
            face_roi = cv2.resize(gray[y1:y2, x1:x2], (200, 200))

            label_id, distance = recognizer.predict(face_roi)
            # Convert distance to confidence %: lower distance = better match
            confidence = max(0, round(100 - distance))
            name = label_names.get(label_id, "Unknown")

            if confidence < CONFIDENCE_THRESHOLD:
                name = "Unknown"

            already = name in marked_today

            if name != "Unknown" and not already:
                if name not in confirm_start:
                    confirm_start[name] = now
                elapsed  = now - confirm_start[name]
                progress = min(elapsed / CONFIRM_SECONDS, 1.0)
                draw_box(frame, x1, y1, x2, y2,
                         name, confidence, "confirming", progress)
                if elapsed >= CONFIRM_SECONDS:
                    df.at[name, today] = "Present"
                    save_excel(df)
                    marked_today.add(name)
                    confirm_start.pop(name, None)
                    t = datetime.now().strftime("%H:%M:%S")
                    print(f"  PRESENT: {name} at {t}")
            elif already:
                confirm_start.pop(name, None)
                draw_box(frame, x1, y1, x2, y2,
                         name, confidence, "marked")
            else:
                draw_box(frame, x1, y1, x2, y2,
                         name, confidence, "unknown")

        draw_hud(frame, today, len(marked_today), len(all_names))
        cv2.imshow("Face Attendance  —  Press Q to quit", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()
    save_excel(df)

    present = len(marked_today)
    total   = len(all_names)
    absent  = [n for n in all_names if n not in marked_today]

    print("\n" + "=" * 50)
    print("  SESSION COMPLETE")
    print("=" * 50)
    print(f"  Present     : {present} / {total}")
    print(f"  Attendance% : {round(present/total*100 if total else 0)}%")
    if absent:
        print(f"  Absent      : {', '.join(absent)}")
    print(f"  Saved -> {EXCEL_FILE}")
    print("=" * 50)
    input("\n  Press Enter to close...")


if __name__ == "__main__":
    os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
    run()
